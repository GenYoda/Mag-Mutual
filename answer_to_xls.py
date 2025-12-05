"""
answers_to_xlsx.py - Convert answers.json to comprehensive Excel report (3 sheets)

Usage:
    python answers_to_xlsx.py forms/answers/answers_20251103_162555.json
    python answers_to_xlsx.py forms/answers/answers_20251103_162555.json --format csv
"""

import json
import sys
from pathlib import Path
from datetime import datetime
import argparse

try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("❌ Required packages not found. Install with:")
    print("   pip install pandas openpyxl")
    sys.exit(1)


class AnswersReportGenerator:
    """Generate comprehensive Excel reports from answers.json"""
    
    def __init__(self, answers_file: str):
        """
        Initialize report generator
        
        Args:
            answers_file: Path to answers.json file
        """
        self.answers_file = Path(answers_file)
        
        if not self.answers_file.exists():
            print(f"❌ File not found: {answers_file}")
            sys.exit(1)
        
        # Load answers
        with open(self.answers_file, 'r', encoding='utf-8') as f:
            self.data = json.load(f)
        
        self.answers = self.data.get('answers', [])
        self.metadata = self.data.get('metadata', {})
    
    def prepare_summary_data(self) -> list:
        """
        Prepare summary sheet data (one row per question)
        
        Returns:
            List of dicts with summary data
        """
        rows = []
        
        for answer in self.answers:
            # Skip skipped questions
            # if answer.get('answer') == 'SKIPPED':
            #     continue
            
            sources = answer.get('sources', [])
            num_sources = len(sources)
            avg_similarity = sum(s.get('similarity', 0) for s in sources) / num_sources if num_sources > 0 else 0
            
            # Get source file list
            source_files = ', '.join(set([s.get('file', 'Unknown') for s in sources]))
            
            row = {
            'Section': answer.get('section_name', ''),
            'Question ID': answer.get('question_id', ''),
            'Question': answer.get('main_question', ''),
            'Question Type': answer.get('question_type', ''),
            'Answer': answer.get('answer', ''),
            'Explanation': answer.get('explanation', ''),
            'Model Reasoning': answer.get('raw_answer', ''),  # ← NEW
            # 'Confidence': f"{answer.get('confidence', 0):.2%}",
            'Source Files': source_files,
            '# Sources': num_sources,
            # 'Avg Similarity': f"{avg_similarity:.1f}%",
            'Page Number': answer.get('page_number', '')
            # 'Used Context': '✓' if answer.get('used_context') else '-'
        }

            
            rows.append(row)
        
        return rows
    
    def prepare_source_detail_data(self) -> list:
        """
        Prepare source details sheet data (one row per source)
        
        Returns:
            List of dicts with source detail data
        """
        rows = []
        
        for answer in self.answers:
            # Skip skipped questions
            # if answer.get('answer') == 'SKIPPED':
            #     continue
            
            q_id = answer.get('question_id', '')
            question = answer.get('main_question', '')
            sources = answer.get('sources', [])
            
            for idx, source in enumerate(sources, 1):
                pages_str = ', '.join(map(str, source.get('pages', [])))
                
                row = {
                    'Question ID': q_id,
                    'Question': question,
                    'Source #': idx,
                    'File': source.get('file', 'Unknown'),
                    'Pages': pages_str,
                    'Similarity': f"{source.get('similarity', 0):.1f}%",
                    'Preview (200 chars)': source.get('chunk_preview', '')
                }
                
                rows.append(row)
        
        return rows
    
    def prepare_full_text_data(self) -> list:
        """
        Prepare full text sheet data (one row per source with full chunk)
        
        Returns:
            List of dicts with full chunk text
        """
        rows = []
        
        for answer in self.answers:
            # Skip skipped questions
            # if answer.get('answer') == 'SKIPPED':
            #     continue
            
            q_id = answer.get('question_id', '')
            question = answer.get('main_question', '')
            sources = answer.get('sources', [])
            
            for idx, source in enumerate(sources, 1):
                pages_str = ', '.join(map(str, source.get('pages', [])))
                
                # Get full chunk text (fallback to preview if not available)
                full_text = source.get('chunk_full', source.get('chunk_preview', ''))
                
                row = {
                    'Question ID': q_id,
                    'Question': question,
                    'Source #': idx,
                    'File': source.get('file', 'Unknown'),
                    'Pages': pages_str,
                    'Similarity': f"{source.get('similarity', 0):.1f}%",
                    'Full Chunk Text': full_text
                }
                
                rows.append(row)
        
        return rows
    
    def save_excel(self, output_file: str = None) -> str:
        """
        Save comprehensive Excel report with 3 sheets
        
        Args:
            output_file: Output file path (default: auto-generated)
            
        Returns:
            Path to output file
        """
        if output_file is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_file = f"answers_report_{timestamp}.xlsx"
        
        output_path = Path(output_file)
        
        # Prepare data for all sheets
        summary_data = self.prepare_summary_data()
        source_detail_data = self.prepare_source_detail_data()
        full_text_data = self.prepare_full_text_data()
        
        if not summary_data:
            print("❌ No data to export")
            return None
        
        # Create Excel writer
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Sheet 1: Summary
            df_summary = pd.DataFrame(summary_data)
            df_summary.to_excel(writer, sheet_name='Summary', index=False)
            
            # Sheet 2: Source Details
            df_source_detail = pd.DataFrame(source_detail_data)
            df_source_detail.to_excel(writer, sheet_name='Source Details', index=False)
            
            # Sheet 3: Full Text
            df_full_text = pd.DataFrame(full_text_data)
            df_full_text.to_excel(writer, sheet_name='Full Chunk Text', index=False)
            
            # Format sheets
            workbook = writer.book
            self._format_summary_sheet(writer.sheets['Summary'], df_summary)
            self._format_source_detail_sheet(writer.sheets['Source Details'], df_source_detail)
            self._format_full_text_sheet(writer.sheets['Full Chunk Text'], df_full_text)
        
        return str(output_path)
    
    def save_csv(self, output_file: str = None) -> str:
        """
        Save report as CSV (summary only)
        
        Args:
            output_file: Output file path (default: auto-generated)
            
        Returns:
            Path to output file
        """
        if output_file is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_file = f"answers_report_{timestamp}.csv"
        
        output_path = Path(output_file)
        
        # Prepare data
        rows = self.prepare_summary_data()
        
        if not rows:
            print("❌ No data to export")
            return None
        
        # Create DataFrame
        df = pd.DataFrame(rows)
        
        # Save CSV
        df.to_csv(output_file, index=False, encoding='utf-8')
        
        return str(output_path)
    
    @staticmethod
    def _format_summary_sheet(worksheet, df):
        """Format summary sheet"""
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Format header
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Column widths
        column_widths = {
            'A': 20,  # Section
            'B': 12,  # Question ID
            'C': 50,  # Question
            'D': 20,  # Question Type
            'E': 25,  # Answer
            'F': 40,  # Explanation
            'G': 50,  # Raw LLM Answer ← NEW
            # 'H': 12,  # Confidence
            'I': 30,  # Source Files
            'J': 10,  # # Sources
            # 'K': 12,  # Avg Similarity
            'L': 12,  # Page Number
            # 'M': 12   # Used Context
        }

        
        for col_letter, width in column_widths.items():
            worksheet.column_dimensions[col_letter].width = width
        
        # Format data rows
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        worksheet.row_dimensions[1].height = 30
        for row in range(2, worksheet.max_row + 1):
            worksheet.row_dimensions[row].height = 40
    
    @staticmethod
    def _format_source_detail_sheet(worksheet, df):
        """Format source detail sheet"""
        header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Format header
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Column widths
        column_widths = {
            'A': 12, 'B': 40, 'C': 10, 'D': 30,
            'E': 15, 'F': 12, 'G': 50
        }
        
        for col_letter, width in column_widths.items():
            worksheet.column_dimensions[col_letter].width = width
        
        # Format data rows
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        worksheet.row_dimensions[1].height = 30
        for row in range(2, worksheet.max_row + 1):
            worksheet.row_dimensions[row].height = 35
    
    @staticmethod
    def _format_full_text_sheet(worksheet, df):
        """Format full text sheet"""
        header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        header_font = Font(bold=True, color="000000", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Format header
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Column widths
        column_widths = {
            'A': 12, 'B': 35, 'C': 10, 'D': 30,
            'E': 15, 'F': 12, 'G': 80  # Wide column for full text
        }
        
        for col_letter, width in column_widths.items():
            worksheet.column_dimensions[col_letter].width = width
        
        # Format data rows
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        worksheet.row_dimensions[1].height = 30
        for row in range(2, worksheet.max_row + 1):
            worksheet.row_dimensions[row].height = 100  # Tall rows for full text
    
    def generate_report(self, output_format: str = 'xlsx', output_file: str = None):
        """
        Generate report in specified format
        
        Args:
            output_format: 'xlsx' or 'csv'
            output_file: Optional output file path
            
        Returns:
            Path to generated file
        """
        print(f"\n{'='*70}")
        print(f"📊 Generating {output_format.upper()} Report")
        print(f"{'='*70}\n")
        
        print(f"📖 Input: {self.answers_file}")
        print(f"📝 Total Answers: {len(self.answers)}")
        
        if output_format.lower() == 'xlsx':
            output_path = self.save_excel(output_file)
            if output_path:
                print(f"\n✅ Excel Report Generated:")
                print(f"   Sheet 1: Summary (one row per question)")
                print(f"   Sheet 2: Source Details (one row per source)")
                print(f"   Sheet 3: Full Chunk Text (complete source text)")
                print(f"   📄 File: {output_path}\n")
        elif output_format.lower() == 'csv':
            output_path = self.save_csv(output_file)
            if output_path:
                print(f"✅ CSV Report saved: {output_path}\n")
        else:
            print(f"❌ Unsupported format: {output_format}")
            return None
        
        return output_path


def main():
    """Main entry point"""
    parser = argparse.ArgumentParser(
        description='Convert answers.json to comprehensive Excel report (3 sheets)'
    )
    parser.add_argument(
        'answers_file',
        help='Path to answers.json file'
    )
    parser.add_argument(
        '--format',
        choices=['xlsx', 'csv'],
        default='xlsx',
        help='Output format (default: xlsx with 3 sheets)'
    )
    parser.add_argument(
        '--output',
        help='Output file path (optional, auto-generated if not provided)'
    )
    
    args = parser.parse_args()
    
    # Generate report
    generator = AnswersReportGenerator(args.answers_file)
    generator.generate_report(
        output_format=args.format,
        output_file=args.output
    )


if __name__ == "__main__":
    main()
